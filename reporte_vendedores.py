#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Reporte Semanal de Vendedores - Temponovo
Ejecuta automáticamente vía GitHub Actions
Lunes y Miércoles a las 11:00 AM hora Chile
"""

import os
import sys
import requests
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from email.mime.base import MIMEBase
import pandas as pd
import xmlrpc.client
import tempfile
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

print("🚀 Iniciando reporte automático...")

# ══════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════

def get_secret(key, default=""):
    """Lee desde variables de entorno (GitHub Secrets)"""
    return os.getenv(key, default)

ODOO3_URL  = get_secret("ODOO3_URL",  "https://odoo.temponovo.cl")
ODOO3_DB   = get_secret("ODOO3_DB",   "temponovo")
ODOO3_USER = get_secret("ODOO3_USER", "admin")
ODOO3_PASS = get_secret("ODOO3_PASS")

SMTP_HOST = get_secret("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(get_secret("SMTP_PORT", "587"))
SMTP_USER = get_secret("SMTP_USER")
SMTP_PASS = get_secret("SMTP_PASS")

# Configuración vendedores

VENDEDORES = [
    {"name": "ALDO CAYAZZO",         "email": "aldocayazzo@hotmail.com",                "id": 5210},
    {"name": "ALEJANDRO STARK",      "email": "starksolla@gmail.com",             "id": 5211},
    {"name": "FRANCISCO BERCZELLER", "email": None,                               "id": 5213},
    {"name": "FRANCISCO CORNEJO", "email": "tatocornejo@yahoo.es", "email2": "tatocornejo1962@gmail.com", "id": 5214},
    {"name": "FREDY ARCHILE",        "email": "fredy@temponovo.cl",              "id": 5216},
    {"name": "MAX LEVY",             "email": "levy.max@gmail.com",               "id": 5217},
    {"name": "OFICINA",              "email": "estrellallanos@temponovo.cl",      "id": 5218},
    {"name": "PEDRO GODOY",          "email": "pedro@aviv.cl", "id": 5219},
    {"name": "Estrella Llanos",      "email": "estrellallanos@temponovo.cl",      "id": None},
]

CC_FIJOS = ["natalia@temponovo.cl", "m.elena@temponovo.cl","estrellallanos@temponovo.cl"]
RESUMEN_EMAILS = ["natalia@temponovo.cl", "admin@temponovo.cl"]

# Nombres cortos para subject del email
NOMBRES_CORTOS = {
    "ALDO CAYAZZO": "Aldo",
    "ALEJANDRO STARK": "Alejandro",
    "FRANCISCO BERCZELLER": "Francisco B.",
    "FRANCISCO CORNEJO": "Francisco C.",
    "FREDY ARCHILE": "Fredy",
    "MAX LEVY": "Max",
    "OFICINA": "Oficina",
    "PEDRO GODOY": "Pedro",
    "Estrella Llanos": "Estrella",
}

# IMPORTANTE: En producción usar False
TEST_MODE = False

# IMPORTANTE: En producción usar False
TEST_MODE = True
TEST_TO   = ["natalia@temponovo.cl", "daniel@temponovo.cl"]

# ── bloque-1b-helpers ──────────────────────────────────────────────────
# ── Helpers many2one ──────────────────────────
def m2o_name(x):
    return x[1] if isinstance(x, list) and len(x) >= 2 else ""

def m2o_id(x):
    return x[0] if isinstance(x, list) and len(x) >= 1 else None

# ── Formato números ───────────────────────────
def format_clp(x):
    """1759125 → '1.759.125'"""
    if pd.isna(x):
        return ""
    try:
        n = int(round(float(x), 0))
    except Exception:
        return ""
    return f"{n:,}".replace(",", ".")

# ── Rangos de fechas ──────────────────────────
def last_7_days_range():
    now   = pd.Timestamp.now()
    end   = now.normalize() + pd.Timedelta(days=1)
    start = now.normalize() - pd.Timedelta(days=7)
    return start, end

# ── Tabla HTML ────────────────────────────────
def df_to_html_table(df: pd.DataFrame) -> str:
    return df.to_html(index=False, escape=False, border=0)

# ── Archivos temporales ───────────────────────
def safe_remove(path: str, retries: int = 6, wait_s: float = 0.3):
    for _ in range(retries):
        try:
            if os.path.exists(path):
                os.remove(path)
            return True
        except PermissionError:
            time.sleep(wait_s)
        except Exception:
            return False
    return False

def safe_filename(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", (s or "").strip())

def export_excel_autowidth(df: pd.DataFrame, path: str, sheet_name: str = "Detalle"):
    df.to_excel(path, index=False, sheet_name=sheet_name)
    wb = load_workbook(path)
    try:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for v in df[col_name].astype(str).head(2000):
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)
        wb.save(path)
    finally:
        wb.close()

# ── CSS base del email ────────────────────────
BASE_CSS = """
<style>
  body  { font-family: Arial, sans-serif; font-size: 13px; color: #111; }
  .wrap { max-width: 980px; }
  h2    { font-size: 17px; margin: 0 0 6px; color: #1a1a1a; }
  h3    { font-size: 14px; margin: 22px 0 6px; color: #333;
          border-bottom: 1px solid #ddd; padding-bottom: 4px; }
  .kpi  { background: #f9f9f9; border-left: 4px solid #4a90d9;
          padding: 8px 14px; margin: 8px 0 16px; }
  .kpi b { font-size: 15px; }
  .lo-que-viene {
      background: #fffbe6; border-left: 4px solid #f5a623;
      padding: 10px 14px; margin: 20px 0 10px; border-radius: 3px;
  }
  table { border-collapse: collapse; width: 100%; margin: 8px 0 16px;
          table-layout: fixed; }
  th, td { border: 1px solid #ddd; padding: 7px 9px;
           text-align: left; vertical-align: top;
           white-space: normal; word-break: break-word; color: #111; }
  th   { background: #2c3e50; color: #ffffff !important;
         font-weight: bold; }
  tr:nth-child(even) td { background: #fafafa; }
  /* tabla cobranza: col cliente ancha, cols monto angostas */
  .tbl-cobr { table-layout: fixed; width: 100%; }
  .tbl-cobr td, .tbl-cobr th { text-align: right; white-space: nowrap; }
  .tbl-cobr td:first-child,
  .tbl-cobr th:first-child { text-align: left; white-space: normal;
                              width: 34%; word-break: break-word; }
  .tbl-cobr td:not(:first-child),
  .tbl-cobr th:not(:first-child) { width: 9%; font-size: 12px; }
  /* celdas vacías → sin color de aviso */
  .tbl-cobr td.vacio { color: #ccc; }
  /* fila vencida >30 días */
  .fila-vencida td { background: #fde8e8 !important; }
  .fila-vencida td.monto-vencido { color: #c0392b; font-weight: bold; }
  .footer { color: #888; font-size: 11px; margin-top: 24px;
            border-top: 1px solid #eee; padding-top: 8px; }
</style>
"""

print("✅ Helpers y CSS cargados")

# ── bloque-2-conexion ──────────────────────────────────────────────────
# ── Conexión XML-RPC (para leer datos) ────────
common = xmlrpc.client.ServerProxy(f"{ODOO3_URL}/xmlrpc/2/common")
uid    = common.authenticate(ODOO3_DB, ODOO3_USER, ODOO3_PASS, {})
models = xmlrpc.client.ServerProxy(f"{ODOO3_URL}/xmlrpc/2/object")
print(f"✅ XML-RPC conectado | uid={uid}")

# ── Conexión HTTP con sesión (para descargar PDFs) ─
sess = requests.Session()
r = sess.post(
    f"{ODOO3_URL}/web/session/authenticate",
    json={"jsonrpc": "2.0", "params": {"db": ODOO3_DB, "login": ODOO3_USER, "password": ODOO3_PASS}},
    timeout=60,
)
r.raise_for_status()
sess_uid = (r.json().get("result") or {}).get("uid")
if not sess_uid:
    raise RuntimeError(f"No autenticó con sesión: {r.json()}")
print(f"✅ Sesión HTTP conectada | uid={sess_uid}")

def jsonrpc_sess(service, method, args):
    payload = {"jsonrpc": "2.0", "method": "call",
               "params": {"service": service, "method": method, "args": args}}
    rr = sess.post(f"{ODOO3_URL}/jsonrpc", json=payload, timeout=120)
    rr.raise_for_status()
    j = rr.json()
    if "error" in j:
        raise RuntimeError(j["error"])
    return j["result"]

# ── bloque-a-ventas ──────────────────────────────────────────────────
start, end = last_7_days_range()
START_STR  = start.strftime("%Y-%m-%d %H:%M:%S")
END_STR    = end.strftime("%Y-%m-%d %H:%M:%S")
print(f"Período: {START_STR} → {END_STR}")

domain_sol = [
    ("order_id.date_order", ">=", START_STR),
    ("order_id.date_order", "<",  END_STR),
    ("order_id.state", "in", ["sale", "done"]),
]
sol_ids = models.execute_kw(
    ODOO3_DB, uid, ODOO3_PASS,
    "sale.order.line", "search",
    [domain_sol], {"limit": 300000}
)
print(f"Líneas encontradas: {len(sol_ids)}")

sols = []
if sol_ids:
    sols = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "sale.order.line", "read",
        [sol_ids],
        {"fields": ["order_id", "order_partner_id", "product_id",
                    "product_uom_qty", "qty_delivered",
                    "price_unit", "price_subtotal", "discount", "display_type"]}
    )

# Prefetch: pedidos (fecha + vendedor)
order_ids = list({m2o_id(s.get("order_id")) for s in sols if m2o_id(s.get("order_id"))})
order_map = {}
if order_ids:
    orders = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "sale.order", "read",
        [order_ids], {"fields": ["date_order", "user_id"]}
    )
    order_map = {o["id"]: o for o in orders}

# Prefetch: clientes (RUT)
partner_ids = list({m2o_id(s.get("order_partner_id")) for s in sols if m2o_id(s.get("order_partner_id"))})
partner_map = {}
if partner_ids:
    partners = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "res.partner", "read",
        [partner_ids], {"fields": ["name", "vat"]}
    )
    partner_map = {p["id"]: p for p in partners}

# Prefetch: productos
product_ids = list({m2o_id(s.get("product_id")) for s in sols if m2o_id(s.get("product_id"))})
product_map = {}
if product_ids:
    products = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "product.product", "read",
        [product_ids], {"fields": ["name", "default_code", "barcode"]}
    )
    product_map = {p["id"]: p for p in products}

# Prefetch: facturas asociadas a cada pedido
# invoice_ids en sale.order → lista de account.move
invoice_map = {}  # order_id → "FAC 083251, FAC 083252"
if order_ids:
    orders_inv = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "sale.order", "read",
        [order_ids], {"fields": ["id", "invoice_ids"]}
    )
    all_inv_ids = []
    order_inv_ids = {}
    for o in orders_inv:
        inv_ids = o.get("invoice_ids") or []
        order_inv_ids[o["id"]] = inv_ids
        all_inv_ids.extend(inv_ids)

    if all_inv_ids:
        inv_names = models.execute_kw(
            ODOO3_DB, uid, ODOO3_PASS,
            "account.move", "read",
            [list(set(all_inv_ids))],
            {"fields": ["id", "name", "state", "move_type"]}
        )
        inv_name_map = {i["id"]: i["name"] for i in inv_names
                        if i.get("state") == "posted"
                        and i.get("move_type") in ("out_invoice", "out_refund")}
        for oid, iids in order_inv_ids.items():
            nombres = [inv_name_map[i] for i in iids if i in inv_name_map]
            invoice_map[oid] = ", ".join(sorted(nombres)) if nombres else ""

# Armar DataFrame
rows = []
for sol in sols:
    if sol.get("display_type"):
        continue
    oid = m2o_id(sol.get("order_id"))
    pid = m2o_id(sol.get("order_partner_id"))
    prd = m2o_id(sol.get("product_id"))
    order   = order_map.get(oid, {})
    partner = partner_map.get(pid, {})
    product = product_map.get(prd, {})
    qty      = float(sol.get("product_uom_qty") or 0)
    price    = float(sol.get("price_unit") or 0)
    disc_pct = float(sol.get("discount") or 0)
    rows.append({
        "Vendedor":             m2o_name(order.get("user_id")),
        "Fecha_pedido":         order.get("date_order", ""),
        "Numero":               m2o_name(sol.get("order_id")),
        "Factura":              invoice_map.get(oid, ""),
        "Cliente":              m2o_name(sol.get("order_partner_id")),
        "RUT":                  (partner.get("vat") or "").strip(),
        "Referencia":           (product.get("default_code") or "").strip(),
        "Pedido":               qty,
        "Entregado":            float(sol.get("qty_delivered") or 0),
        "Descuento_%":          disc_pct,
        "Precio_unit_con_desc": price * (1 - disc_pct / 100),
        "Total":                float(sol.get("price_subtotal") or 0),
    })

df_ventas = pd.DataFrame(rows)
df_ventas["Fecha_pedido"] = pd.to_datetime(df_ventas["Fecha_pedido"], errors="coerce")
for c in ["Pedido", "Entregado", "Descuento_%", "Precio_unit_con_desc", "Total"]:
    df_ventas[c] = pd.to_numeric(df_ventas[c], errors="coerce").round(0).astype("Int64")

print(f"\n✅ df_ventas: {len(df_ventas)} líneas")
df_ventas.head(3)

# ── bloque-b-diferencias ──────────────────────────────────────────────────
dv_all = df_ventas.copy()
dv_all["Entregado"] = pd.to_numeric(dv_all["Entregado"], errors="coerce").fillna(0)
dv_all["Pedido"]    = pd.to_numeric(dv_all["Pedido"],    errors="coerce").fillna(0)

flags = dv_all.assign(
    hay_cero  = (dv_all["Entregado"] == 0),
    hay_mayor = (dv_all["Entregado"]  > 0),
).groupby("Numero", as_index=False).agg(
    tiene_cero  = ("hay_cero",  "any"),
    tiene_mayor = ("hay_mayor", "any"),
)

nums_inc = flags.loc[flags["tiene_cero"] & flags["tiene_mayor"], "Numero"]

df_diferencias = dv_all[
    dv_all["Numero"].isin(nums_inc) & (dv_all["Entregado"] == 0)
].copy()
df_diferencias["Pendiente"] = (
    df_diferencias["Pedido"] - df_diferencias["Entregado"]
).clip(lower=0).astype("Int64")

print(f"✅ Pedidos con diferencias: {df_diferencias['Numero'].nunique()}")
print(f"   Líneas sin entregar    : {len(df_diferencias)}")
df_diferencias[["Vendedor", "Numero", "Cliente", "Referencia",
                "Pedido", "Entregado", "Pendiente"]].head(5)

# ── bloque-c-pendientes ──────────────────────────────────────────────────
print("Cargando pedidos de venta...")
so_ids = models.execute_kw(
    ODOO3_DB, uid, ODOO3_PASS,
    "sale.order", "search",
    [[]], {"limit": 300000}
)
orders_raw = []
if so_ids:
    orders_raw = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "sale.order", "read",
        [so_ids],
        {"fields": ["id", "date_order", "name", "partner_id",
                    "user_id", "tag_ids", "state",
                    "invoice_status", "amount_total"]}
    )

# Etiquetas
all_tag_ids = sorted({tid for o in orders_raw for tid in (o.get("tag_ids") or [])})
tag_map = {}
if all_tag_ids:
    tags = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "crm.tag", "read", [all_tag_ids], {"fields": ["name"]}
    )
    tag_map = {t["id"]: (t.get("name") or "") for t in tags}

# Líneas → categorías padre
order_map_c = {o["id"]: o for o in orders_raw}
sol_ids_c = models.execute_kw(
    ODOO3_DB, uid, ODOO3_PASS,
    "sale.order.line", "search",
    [[('order_id', 'in', list(order_map_c.keys())), ('display_type', '=', False)]],
    {"limit": 300000}
)
sols_c = []
if sol_ids_c:
    sols_c = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "sale.order.line", "read",
        [sol_ids_c], {"fields": ["order_id", "product_id", "product_uom_qty"]}
    )

# producto → template → categoría → padre
prod_ids_c = sorted({m2o_id(s.get("product_id")) for s in sols_c if m2o_id(s.get("product_id"))})
prod_to_tmpl = {}
if prod_ids_c:
    prods = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "product.product", "read", [prod_ids_c], {"fields": ["id", "product_tmpl_id"]}
    )
    prod_to_tmpl = {p["id"]: m2o_id(p.get("product_tmpl_id")) for p in prods}

tmpl_to_categ = {}
tmpl_ids_c = sorted({t for t in prod_to_tmpl.values() if t})
if tmpl_ids_c:
    tmpls = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "product.template", "read", [tmpl_ids_c], {"fields": ["id", "categ_id"]}
    )
    tmpl_to_categ = {t["id"]: m2o_id(t.get("categ_id")) for t in tmpls}

cat_map = {}
categ_ids_c = sorted({c for c in tmpl_to_categ.values() if c})
if categ_ids_c:
    cats = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "product.category", "read", [categ_ids_c], {"fields": ["id", "name", "parent_id"]}
    )
    parent_ids_needed = set()
    for c in cats:
        pid = m2o_id(c.get("parent_id"))
        if pid:
            parent_ids_needed.add(pid)
        cat_map[c["id"]] = {"name": c.get("name") or "", "parent_id": pid}
    missing = sorted([p for p in parent_ids_needed if p not in cat_map])
    if missing:
        pcats = models.execute_kw(
            ODOO3_DB, uid, ODOO3_PASS,
            "product.category", "read", [missing], {"fields": ["id", "name", "parent_id"]}
        )
        for c in pcats:
            cat_map[c["id"]] = {"name": c.get("name") or "", "parent_id": m2o_id(c.get("parent_id"))}

def parent_cat_name(categ_id):
    if not categ_id or categ_id not in cat_map:
        return "Sin categoría"
    pid = cat_map[categ_id].get("parent_id")
    if pid and pid in cat_map and cat_map[pid].get("name"):
        return cat_map[pid]["name"]
    return cat_map[categ_id].get("name") or "Sin categoría"

order_cat_qty = {}
for sol in sols_c:
    oid = m2o_id(sol.get("order_id"))
    pid = m2o_id(sol.get("product_id"))
    qty = float(sol.get("product_uom_qty") or 0)
    if not oid or not pid:
        continue
    tmpl_id  = prod_to_tmpl.get(pid)
    categ_id = tmpl_to_categ.get(tmpl_id) if tmpl_id else None
    cat_name = parent_cat_name(categ_id)
    order_cat_qty.setdefault(oid, {})
    order_cat_qty[oid][cat_name] = order_cat_qty[oid].get(cat_name, 0) + qty

def format_cats(d):
    if not d:
        return ""
    parts = []
    for k in sorted(d.keys()):
        v = d[k]
        v_out = int(round(v)) if abs(v - round(v)) < 1e-9 else round(v, 2)
        parts.append(f"{k}: {v_out}")
    return ", ".join(parts)

state_map   = {"draft": "Borrador", "sent": "Enviado", "sale": "Confirmado",
               "done": "Cerrado", "cancel": "Cancelado"}
invoice_map = {"no": "Nada que facturar", "to invoice": "Por facturar", "invoiced": "Facturado"}

rows_c = []
for o in orders_raw:
    tag_names = [tag_map.get(tid, "") for tid in (o.get("tag_ids") or [])]
    tag_names = [t for t in tag_names if t]
    rows_c.append({
        "Fecha_pedido":       o.get("date_order") or "",
        "Numero":             o.get("name") or "",
        "Cliente":            m2o_name(o.get("partner_id")),
        "Vendedor":           m2o_name(o.get("user_id")),
        "Etiquetas":          ", ".join(tag_names),
        "Estado":             state_map.get(o.get("state"), o.get("state")),
        "Estado_facturacion": invoice_map.get(o.get("invoice_status"), o.get("invoice_status")),
        "Total_pedido":       float(o.get("amount_total") or 0),
        "Categorias":         format_cats(order_cat_qty.get(o.get("id"), {})),
    })

df_pend_raw = pd.DataFrame(rows_c)
df_pend_raw["Fecha_pedido"] = pd.to_datetime(df_pend_raw["Fecha_pedido"], errors="coerce")
hoy = pd.Timestamp.now().normalize()

df_pendientes = df_pend_raw[
    df_pend_raw["Estado"].isin(["Borrador", "Enviado"]) &
    (df_pend_raw["Fecha_pedido"] < hoy)
].copy()

print(f"✅ df_pendientes: {len(df_pendientes)} pedidos pendientes de pago")
df_pendientes[["Fecha_pedido", "Numero", "Cliente", "Vendedor",
               "Total_pedido", "Categorias"]].head(5)

# ── bloque-d-cobranza ──────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, KeepTogether)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import io

print("Cargando saldos por cobrar (facturas + asientos de apertura)...")
hoy_dt = pd.Timestamp.now().normalize()

# ══════════════════════════════════════════════════════════════
# ESTRATEGIA: account.move.line filtrando por cuentas por cobrar
# Captura TODO: FAC, ND, NC y asientos de apertura (APER/).
# Compatibilidad Odoo 16/17: buscamos las cuentas receivable
# primero y luego filtramos las líneas por esas cuentas.
# ══════════════════════════════════════════════════════════════

# ── 1a) Obtener IDs de cuentas por cobrar ─────────────────
# La cuenta es A110402 **CLIENTES (visible en el APER/)
# Buscamos por código A110402 y cualquier variante
receivable_ids = []

# Intento 1: código exacto A110402
try:
    receivable_ids = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "account.account", "search",
        [[("code", "=like", "A110402%")]],
        {"limit": 200}
    )
except Exception:
    pass

# Intento 2: account_type estándar Odoo 16+
if not receivable_ids:
    try:
        receivable_ids = models.execute_kw(
            ODOO3_DB, uid, ODOO3_PASS,
            "account.account", "search",
            [[("account_type", "=", "asset_receivable")]],
            {"limit": 200}
        )
    except Exception:
        pass

# Intento 3: nombre contiene CLIENTES
if not receivable_ids:
    try:
        receivable_ids = models.execute_kw(
            ODOO3_DB, uid, ODOO3_PASS,
            "account.account", "search",
            [[("name", "ilike", "CLIENTES")]],
            {"limit": 200}
        )
    except Exception:
        pass

print(f"Cuentas por cobrar encontradas: {len(receivable_ids)}")
if receivable_ids:
    # Mostrar nombres para confirmar que son las correctas
    acc_names = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "account.account", "read",
        [receivable_ids], {"fields": ["code", "name"]}
    )
    for a in acc_names[:5]:
        print(f"   {a['code']}  {a['name']}")
else:
    raise RuntimeError("No se encontraron cuentas por cobrar. Revisar plan de cuentas.")

# ── 1b) Líneas no reconciliadas de esas cuentas ────────────
# Para capturar tanto facturas (FAC) como asientos de apertura (APER)
# no filtramos por amount_residual > 0 en el search (ese campo
# en account.move.line se comporta diferente según move_type).
# En cambio traemos todas las líneas no reconciliadas y filtramos
# por debit > 0 (líneas de cargo al cliente) en Python.
aml_domain = [
    ("account_id",        "in",  receivable_ids),
    ("move_id.state",     "=",   "posted"),
    ("partner_id",        "!=",  False),
    ("full_reconcile_id", "=",   False),   # no totalmente conciliada
]
aml_ids = models.execute_kw(
    ODOO3_DB, uid, ODOO3_PASS,
    "account.move.line", "search",
    [aml_domain], {"limit": 200000}
)
print(f"Líneas por cobrar candidatas: {len(aml_ids)}")

aml_list = []
if aml_ids:
    aml_list = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "account.move.line", "read",
        [aml_ids],
        {"fields": ["move_id", "partner_id", "date_maturity",
                    "date", "debit", "credit", "amount_residual",
                    "full_reconcile_id", "name"]}
    )
    # Filtrar: solo líneas con saldo real pendiente
    # debit > credit significa que el cliente nos debe
    # full_reconcile_id = False significa no totalmente conciliada
    aml_list = [
        l for l in aml_list
        if not l.get("full_reconcile_id")                          # no conciliada totalmente
        and (float(l.get("amount_residual") or 0) > 0             # tiene saldo
             or float(l.get("debit") or 0) > float(l.get("credit") or 0))  # o debit > credit
    ]
    print(f"Líneas por cobrar con saldo: {len(aml_list)}")

# ── 2) Prefetch datos del asiento (nombre, vendedor) ───────
move_ids_d = list({m2o_id(l.get("move_id")) for l in aml_list if m2o_id(l.get("move_id"))})
move_map   = {}
if move_ids_d:
    moves = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "account.move", "read",
        [move_ids_d],
        {"fields": ["id", "name", "ref", "invoice_user_id", "move_type",
                    "invoice_date", "invoice_date_due", "amount_residual"]}
    )
    move_map = {m["id"]: m for m in moves}

# ── 3) Prefetch ciudad de los partners ─────────────────────
partner_ids_d = list({m2o_id(l.get("partner_id")) for l in aml_list
                      if m2o_id(l.get("partner_id"))})
city_map = {}
if partner_ids_d:
    partners_d = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "res.partner", "read",
        [partner_ids_d], {"fields": ["id", "city"]}
    )
    city_map = {p["id"]: (p.get("city") or "").strip() for p in partners_d}

# ── 4) Zonas Norte / Centro / Sur ──────────────────────────
CIUDADES = {
    # ── NORTE ─────────────────────────────────────────────
    "arica":              (0,  0), "parinacota":         (0,  1),
    "iquique":            (0,  2), "alto hospicio":      (0,  3),
    "pozo almonte":       (0,  4), "tocopilla":          (0,  5),
    "calama":             (0,  6), "antofagasta":        (0,  7),
    "mejillones":         (0,  8), "taltal":             (0,  9),
    "chañaral":           (0, 10), "chanarral":          (0, 10),
    "copiapó":            (0, 11), "copiapo":            (0, 11),
    "vallenar":           (0, 12), "huasco":             (0, 13),
    "freirina":           (0, 14),
    "la serena":          (0, 15), "coquimbo":           (0, 16),
    "ovalle":             (0, 17), "combarbala":         (0, 18),
    "illapel":            (0, 19), "salamanca":          (0, 20),
    "los vilos":          (0, 21),
    # ── CENTRO ────────────────────────────────────────────
    # V Región
    "valparaíso":         (1,  0), "valparaiso":         (1,  0),
    "viña del mar":       (1,  1), "vina del mar":       (1,  1),
    "quillota":           (1,  2), "la calera":          (1,  3),
    "hijuelas":           (1,  4), "llaillay":           (1,  5),
    "san felipe":         (1,  6), "los andes":          (1,  7),
    "cabildo":            (1,  8), "la ligua":           (1,  9),
    "casablanca":         (1, 10), "san antonio":        (1, 11),
    "villa alemana":      (1, 12), "quilpué":            (1, 13),
    "quilpue":            (1, 13), "limache":            (1, 14),
    # RM
    "santiago":           (1, 15), "providencia":        (1, 16),
    "las condes":         (1, 17), "vitacura":           (1, 18),
    "ñuñoa":              (1, 19), "nunoa":              (1, 19),
    "maipú":              (1, 20), "maipu":              (1, 20),
    "pudahuel":           (1, 21), "quilicura":          (1, 22),
    "recoleta":           (1, 23), "independencia":      (1, 24),
    "la florida":         (1, 25), "puente alto":        (1, 26),
    "san bernardo":       (1, 27), "buin":               (1, 28),
    "paine":              (1, 29), "melipilla":          (1, 30),
    "talagante":          (1, 31), "peñaflor":           (1, 32),
    "penaflor":           (1, 32), "colina":             (1, 33),
    "lampa":              (1, 34),
    # VI–VII
    "rancagua":           (1, 35), "san vicente tt":     (1, 36),
    "san vicente":        (1, 36), "peumo":              (1, 37),
    "san fernando":       (1, 38), "santa cruz":         (1, 39),
    "rengo":              (1, 40), "hualañe":            (1, 41),
    "hualane":            (1, 41),
    "curicó":             (1, 42), "curico":             (1, 42),
    "molina":             (1, 43), "talca":              (1, 44),
    "maule":              (1, 45), "san javier":         (1, 46),
    "linares":            (1, 47), "parral":             (1, 48),
    "cauquenes":          (1, 49), "constitución":       (1, 50),
    "constitucion":       (1, 50),
    # VIII norte
    "chillán":            (1, 51), "chillan":            (1, 51),
    "chillán viejo":      (1, 52), "chillan viejo":      (1, 52),
    "san carlos":         (1, 53), "cabrero":            (1, 54),
    # ── SUR ───────────────────────────────────────────────
    "los angeles":        (2,  0), "laja":               (2,  1),
    "angol":              (2,  2), "traiguen":           (2,  3),
    "traiguén":           (2,  3), "curacautin":         (2,  4),
    "curacautín":         (2,  4),
    "concepción":         (2,  5), "concepcion":         (2,  5),
    "talcahuano":         (2,  6), "tome":               (2,  7),
    "tomé":               (2,  7), "hualpén":            (2,  8),
    "hualpen":            (2,  8), "coronel":            (2,  9),
    "lota":               (2, 10), "lebu":               (2, 11),
    "cañete":             (2, 12), "canete":             (2, 12),
    "arauco":             (2, 13), "curanilahue":        (2, 14),
    "temuco":             (2, 15), "lautaro":            (2, 16),
    "victoria":           (2, 17), "nueva imperial":     (2, 18),
    "cholchol":           (2, 19), "cholchól":           (2, 19),
    "tolten":             (2, 20), "toltén":             (2, 20),
    "villarrica":         (2, 21), "villarica":          (2, 21),
    "pucón":              (2, 22), "pucon":              (2, 22),
    "loncoche":           (2, 23), "pitrufquén":         (2, 24),
    "pitrufquen":         (2, 24),
    "valdivia":           (2, 25), "la unión":           (2, 26),
    "la union":           (2, 26), "rio bueno":          (2, 27),
    "osorno":             (2, 28), "puerto octay":       (2, 29),
    "frutillar":          (2, 30),
    "puerto montt":       (2, 31), "puerto varas":       (2, 32),
    "castro":             (2, 33), "ancud":              (2, 34),
    "quemchi":            (2, 35), "quellón":            (2, 36),
    "quellon":            (2, 36), "futaleufu":          (2, 37),
    "futaleufú":          (2, 37), "futalefú":           (2, 37),
    "coyhaique":          (2, 38), "coihaique":          (2, 38),
    "puerto aysen":       (2, 39), "puerto aysén":       (2, 39),
    "puerto cisnes":      (2, 40), "cochrane":           (2, 41),
    "punta arenas":       (2, 42), "punta arena":        (2, 42),
    "puerto natales":     (2, 43), "porvenir":           (2, 44),
}

ZONA_NOMBRES = {0: "🌵  NORTE", 1: "🏙️  CENTRO", 2: "🌲  SUR"}
ZONA_COLORS  = {
    0: colors.HexColor("#1a6b3c"),
    1: colors.HexColor("#1a4a7a"),
    2: colors.HexColor("#6b2d1a"),
}

def get_zona_orden(city: str):
    if not city:
        return (3, "zzz")
    cl = city.lower().strip()
    if cl in CIUDADES:
        return CIUDADES[cl]
    for key, val in CIUDADES.items():
        if cl.startswith(key) or key.startswith(cl):
            return val
    return (3, cl)

# ── 5) Bucket de antigüedad ────────────────────────────────
def bucket(dias):
    if dias <= 0:    return "A_la_fecha"
    elif dias <= 30:  return "d_1_30"
    elif dias <= 60:  return "d_31_60"
    elif dias <= 90:  return "d_61_90"
    elif dias <= 120: return "d_91_120"
    else:             return "Antiguos"

# ── 6) Mapa partner → vendedor ────────────────────────────
# Fuente 1: historial de facturas (más confiable — último vendedor que facturó)
# Fuente 2: user_id del partner (vendedor asignado en la ficha del cliente)
partner_vendor_map = {}

# Fuente 2 primero (base): user_id de todos los partners con saldo
all_pids = list({m2o_id(l.get("partner_id")) for l in aml_list
                 if m2o_id(l.get("partner_id"))})
if all_pids:
    # Leer en chunks de 500 para no sobrecargar
    for i in range(0, len(all_pids), 500):
        chunk = all_pids[i:i+500]
        pdata = models.execute_kw(
            ODOO3_DB, uid, ODOO3_PASS,
            "res.partner", "read",
            [chunk], {"fields": ["id", "user_id"]}
        )
        for p in pdata:
            vend = m2o_name(p.get("user_id"))
            if p["id"] and vend:
                partner_vendor_map[p["id"]] = vend
print(f"  Desde ficha partner (user_id): {len(partner_vendor_map)}")

# Fuente 1 encima: facturas reales sobreescriben (más reciente gana)
inv_factura_ids = models.execute_kw(
    ODOO3_DB, uid, ODOO3_PASS,
    "account.move", "search",
    [[("move_type", "in", ["out_invoice", "out_refund"]),
      ("state", "=", "posted"),
      ("invoice_user_id", "!=", False)]],
    {"limit": 50000}
)
if inv_factura_ids:
    fac_vend = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "account.move", "read",
        [inv_factura_ids],
        {"fields": ["partner_id", "invoice_user_id"]}
    )
    for f in fac_vend:
        pid_f = m2o_id(f.get("partner_id"))
        vend  = m2o_name(f.get("invoice_user_id"))
        if pid_f and vend:
            partner_vendor_map[pid_f] = vend
print(f"  Total mapa partner→vendedor:   {len(partner_vendor_map)}")

# ── 7) Armar filas ─────────────────────────────────────────
rows_d = []
for line in aml_list:
    pid      = m2o_id(line.get("partner_id"))
    move_id  = m2o_id(line.get("move_id"))
    move     = move_map.get(move_id, {})
    ciudad   = city_map.get(pid, "")
    zona_idx, zona_ord = get_zona_orden(ciudad)

    # Fecha de vencimiento: primero date_maturity de la línea,
    # luego invoice_date_due del asiento
    fecha_venc = pd.to_datetime(
        line.get("date_maturity") or move.get("invoice_date_due") or
        line.get("date") or "", errors="coerce"
    )
    # Fecha del documento
    fecha_doc = (move.get("invoice_date") or move.get("date") or
                 line.get("date") or "")

    # Saldo de la línea individual (para el detalle del PDF)
    saldo_linea = float(line.get("amount_residual") or 0)
    if saldo_linea <= 0:
        continue   # ignorar líneas ya pagadas o con saldo negativo
    
    # Saldo del asiento completo (para el resumen pivot)
    saldo_asiento = float(move.get("amount_residual") or 0)
    
    # Si el saldo del asiento es negativo, significa que LE DEBEMOS al cliente (no nos debe)
    # Excluir estos casos completamente
    if saldo_asiento < 0:
        continue
    
    dias = int((hoy_dt - fecha_venc).days) if pd.notna(fecha_venc) else 0

    # Excluir cheques en cartera (ya entregados por el cliente)
    move_ref = (move.get("ref") or "").lower()
    if "cheque" in move_ref and "cartera" in move_ref:
        continue

    # Vendedor: desde el asiento
    vendedor = m2o_name(move.get("invoice_user_id")) or ""
    # Si el asiento no tiene vendedor (APER, entries), inferir del historial
    if not vendedor:
        vendedor = partner_vendor_map.get(pid, "")

    # Nombre del documento
    move_name = move.get("name") or ""
    line_name  = (line.get("name") or "").strip()
    # Key único para groupby/pivot: move_id + line_id (evita duplicados)
    line_id   = line.get("id") or 0
    factura_key = f"{move_name}||{line_id}"
    # Label visual para el PDF
    if move_name.startswith("APER/") and line_name and line_name not in ("/", ""):
        doc_label = f"{move_name} – {line_name}"
    else:
        doc_label = move_name

    rows_d.append({
        "Factura":       factura_key,   # key único para pivot
        "Doc_label":     doc_label,     # texto visual para PDF
        "Move_name":     move_name,
        "Cliente":       m2o_name(line.get("partner_id")),
        "Ciudad":        ciudad,
        "Zona_idx":      zona_idx,
        "Zona_ord":      zona_ord,
        "Vendedor":      vendedor,
        "Fecha_factura": fecha_doc,
        "Fecha_venc":    fecha_venc,
        "Dias_vencido":  dias,
        "Bucket":        bucket(dias),
        "Saldo":         saldo_linea,        # para el PDF (detalle)
        "Saldo_asiento": saldo_asiento,      # para el pivot (resumen)
    })

df_cobr_raw = pd.DataFrame(rows_d)
# Agrupar líneas del mismo asiento+cliente (puede haber múltiples líneas receivable)
df_cobr_raw = (
    df_cobr_raw.groupby(
        ["Factura","Doc_label","Move_name","Cliente","Ciudad","Zona_idx","Zona_ord",
         "Vendedor","Fecha_factura","Fecha_venc","Dias_vencido","Bucket"],
        as_index=False
    ).agg(Saldo=("Saldo","sum"), Saldo_asiento=("Saldo_asiento","first"))
)
# Filtrar solo asientos con saldo pendiente (pero mantener todas las líneas del asiento para el PDF)
df_cobr_raw = df_cobr_raw[df_cobr_raw["Saldo_asiento"] > 0].copy()

# Mostrar ciudades sin zona para ajustar
sin_zona = df_cobr_raw[df_cobr_raw["Zona_idx"]==3]["Ciudad"].value_counts()
if not sin_zona.empty:
    print("⚠️  Ciudades sin zona (agregar al mapa si hace falta):")
    for c, n in sin_zona.items():
        print(f"   {c!r:30s}  ({n} registros)")

if df_cobr_raw.empty:
    print("Sin saldos pendientes.")
    df_cobranza   = pd.DataFrame()
    pdfs_cobranza = {}
else:
    # ── 7) Pivot resumen (para email/tabla) ───────────────────────
    # Usa Saldo_asiento que es el residual del asiento padre (correcto)
    meta_cli = df_cobr_raw.groupby(["Vendedor","Cliente"]).agg(
        Ciudad  =("Ciudad",   "first"),
        Zona_idx=("Zona_idx", "first"),
        Zona_ord=("Zona_ord", "first"),
    ).reset_index()

    pivot = df_cobr_raw.pivot_table(
        index=["Vendedor","Cliente"], columns="Bucket",
        values="Saldo_asiento", aggfunc="sum", fill_value=0
    ).reset_index()

    for col in ["A_la_fecha","d_1_30","d_31_60","d_61_90","d_91_120","Antiguos"]:
        if col not in pivot.columns:
            pivot[col] = 0

    # Renombrar primero
    pivot = pivot.rename(columns={
        "A_la_fecha":"A la fecha","d_1_30":"1-30",
        "d_31_60":"31-60","d_61_90":"61-90","d_91_120":"91-120"
    })
    # Crear columna >30 (consolidada)
    pivot[">30"] = pivot[["31-60","61-90","91-120","Antiguos"]].sum(axis=1)
    # Total = suma de las columnas QUE SE MUESTRAN
    pivot["Total"] = pivot[["A la fecha","1-30",">30"]].sum(axis=1)
    pivot = pivot[pivot["Total"] > 0].copy()
    pivot = pivot.merge(meta_cli, on=["Vendedor","Cliente"], how="left")
    pivot["Ciudad"]   = pivot["Ciudad"].fillna("")
    pivot["Zona_idx"] = pivot["Zona_idx"].fillna(3).astype(int)

    COLS_MONTO  = ["A la fecha","1-30",">30","Total"]
    df_cobranza = pivot[["Vendedor","Cliente","Ciudad","Zona_idx"] + COLS_MONTO].copy()

    # ── 8) PDF por vendedor con zonas ──────────────────────
    COLS_VENCIDO = ["31-60","61-90","91-120","Antiguos"]
    styles  = getSampleStyleSheet()
    st_h1   = ParagraphStyle("h1",  parent=styles["Heading1"], fontSize=13, spaceAfter=4)
    st_sub  = ParagraphStyle("sub", parent=styles["Normal"],   fontSize=9,
                              textColor=colors.gray, spaceAfter=8)
    st_cell = ParagraphStyle("cell",parent=styles["Normal"],   fontSize=8)
    st_zona = ParagraphStyle("zona",parent=styles["Normal"],   fontSize=10,
                              fontName="Helvetica-Bold", textColor=colors.white)

    HDR_COLOR  = colors.HexColor("#2c3e50")
    HDR_TXT    = colors.white
    CLI_BG     = colors.HexColor("#dce8f5")
    VENC_BG    = colors.HexColor("#fde8e8")
    VENC_TXT   = colors.HexColor("#c0392b")
    ALT_BG     = colors.HexColor("#f8f9fa")
    GRID_COLOR = colors.HexColor("#cccccc")

    COL_W   = [10*cm, 3*cm, 3*cm, 3*cm, 3.2*cm]
    HEADERS = ["Cliente / Documento", "A la fecha", "1-30", ">30 días", "Total"]

    def fmt(x):
        try:
            v = float(x)
            return "" if v == 0 else "$ " + f"{int(round(v)):,}".replace(",",".")
        except:
            return ""

    def make_zona_sep(nombre, color):
        data = [[Paragraph(nombre, st_zona), "", "", "", ""]]
        tbl  = Table(data, colWidths=COL_W)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0,0),(-1,0), color),
            ("SPAN",         (0,0),(-1,0)),
            ("TOPPADDING",   (0,0),(-1,0), 6),
            ("BOTTOMPADDING",(0,0),(-1,0), 6),
            ("LEFTPADDING",  (0,0),(-1,0), 10),
        ]))
        return tbl

    def build_pdf_cobranza(df_vend_res, df_raw_vend, nombre_vendedor):
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=1.2*cm, rightMargin=1.2*cm,
                                topMargin=1.2*cm,  bottomMargin=1.2*cm)
        story = []
        story.append(Paragraph(f"Cobranza pendiente – {nombre_vendedor}", st_h1))
        story.append(Paragraph(
            f"Al {hoy_dt.strftime('%d/%m/%Y')}  |  Saldos No Pagados", st_sub))

        # Ordenar clientes: zona → orden_interno → nombre
        base = df_raw_vend[["Cliente","Ciudad","Zona_idx","Zona_ord"]]                   .drop_duplicates("Cliente").copy()
        base = base.sort_values(["Zona_idx","Zona_ord","Cliente"])

        zona_actual = None

        for _, cli_row in base.iterrows():
            cliente  = cli_row["Cliente"]
            zona_idx = int(cli_row["Zona_idx"])
            ciudad   = cli_row["Ciudad"]

            # Separador cuando cambia la zona
            if zona_idx != zona_actual:
                if zona_actual is not None:
                    story.append(Spacer(1, 0.4*cm))
                nombre_zona = ZONA_NOMBRES.get(zona_idx, "📍  SIN ZONA")
                color_zona  = ZONA_COLORS.get(zona_idx, colors.HexColor("#555555"))
                story.append(make_zona_sep(nombre_zona, color_zona))
                story.append(Spacer(1, 0.15*cm))
                zona_actual = zona_idx

            df_cli  = df_raw_vend[df_raw_vend["Cliente"]==cliente].sort_values("Fecha_factura")
            row_res = df_vend_res[df_vend_res["Cliente"]==cliente]

            tiene_mayor30 = False
            for col in COLS_VENCIDO:
                val = float(row_res[col].values[0]) if (
                    not row_res.empty and col in row_res.columns) else 0
                if val > 0:
                    tiene_mayor30 = True; break

            ciudad_tag = (f" <font size='7' color='#888888'>{ciudad}</font>"
                          if ciudad else "")
            cli_label  = Paragraph(f"<b>{cliente}</b>{ciudad_tag}", st_cell)

            if not row_res.empty:
                r = row_res.iloc[0]
                fila_cli = [cli_label,
                            fmt(r.get("A la fecha",0)),
                            fmt(r.get("1-30",      0)),
                            fmt(r.get(">30",       0)),
                            fmt(r.get("Total",     0))]
            else:
                fila_cli = [cli_label, "", "", "", ""]

            filas_fac = []
            for _, inv in df_cli.iterrows():
                fec     = pd.to_datetime(inv["Fecha_factura"], errors="coerce")
                fec_str = fec.strftime("%d/%m/%Y") if pd.notna(fec) else ""
                b, s    = inv["Bucket"], inv["Saldo"]
                # Usar Doc_label (con detalle APER) para mostrar en PDF
                nombre_doc = inv.get("Doc_label") or inv["Factura"].split("||")[0] or "—"
                # Para APER: la fecha ya está en el label, no repetir
                fecha_display = "" if inv["Factura"].startswith("APER/") else f"  ({fec_str})"
                filas_fac.append([
                    Paragraph(f"  {nombre_doc}{fecha_display}", st_cell),
                    fmt(s) if b == "A_la_fecha" else "",
                    fmt(s) if b == "d_1_30"     else "",
                    fmt(s) if b in ("d_31_60","d_61_90","d_91_120","Antiguos") else "",
                    fmt(s),
                ])

            data_grp = [HEADERS, fila_cli] + filas_fac
            tbl = Table(data_grp, colWidths=COL_W, repeatRows=1)

            cmds = [
                ("BACKGROUND",    (0,0),(-1,0),  HDR_COLOR),
                ("TEXTCOLOR",     (0,0),(-1,0),  HDR_TXT),
                ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
                ("BACKGROUND",    (0,1),(-1,1),  CLI_BG),
                ("FONTNAME",      (0,1),(-1,1),  "Helvetica-Bold"),
                ("ROWBACKGROUNDS",(0,2),(-1,-1), [colors.white, ALT_BG]),
                ("GRID",          (0,0),(-1,-1), 0.4, GRID_COLOR),
                ("ALIGN",         (1,0),(-1,-1), "RIGHT"),
                ("ALIGN",         (0,0),(0,-1),  "LEFT"),
                ("FONTSIZE",      (0,0),(-1,-1), 8),
                ("TOPPADDING",    (0,0),(-1,-1), 3),
                ("BOTTOMPADDING", (0,0),(-1,-1), 3),
            ]
            if tiene_mayor30:
                cmds += [("BACKGROUND",(0,1),(-1,1),VENC_BG),
                         ("TEXTCOLOR", (2,1),(-1,1),VENC_TXT)]
            for fi, (_, inv) in enumerate(df_cli.iterrows(), start=2):
                if inv["Bucket"] in ("d_31_60","d_61_90","d_91_120","Antiguos"):
                    cmds += [("BACKGROUND",(0,fi),(-1,fi),VENC_BG),
                             ("TEXTCOLOR", (1,fi),(-1,fi),VENC_TXT)]
            tbl.setStyle(TableStyle(cmds))
            story.append(KeepTogether([tbl, Spacer(1, 0.2*cm)]))

        doc.build(story)
        return buf.getvalue()

    # ── 9) Generar un PDF por vendedor ─────────────────────
    # Filtrar df_cobr_raw para incluir solo clientes con saldo pendiente (que están en df_cobranza)
    clientes_con_saldo = set(df_cobranza[["Vendedor", "Cliente"]].apply(tuple, axis=1))
    df_cobr_raw_filtrado = df_cobr_raw[
        df_cobr_raw[["Vendedor", "Cliente"]].apply(tuple, axis=1).isin(clientes_con_saldo)
    ].copy()
    
    pdfs_cobranza = {}
    for v in VENDEDORES:
        nombre   = v["name"]
        df_v_res = df_cobranza[df_cobranza["Vendedor"].str.strip()==nombre.strip()].copy()
        df_v_raw = df_cobr_raw_filtrado[df_cobr_raw_filtrado["Vendedor"].str.strip()==nombre.strip()].copy()
        if df_v_res.empty:
            continue
        try:
            pdfs_cobranza[nombre] = build_pdf_cobranza(df_v_res, df_v_raw, nombre)
            print(f"  ✅ PDF {nombre} → {len(pdfs_cobranza[nombre]):,} bytes")
        except Exception as e:
            print(f"  ⚠️  {nombre}: {e}")

    # ── PDF GENERAL: todos los vendedores juntos ──────────────
    try:
        pdf_general = build_pdf_cobranza(
            df_cobranza,           # resumen completo (solo con saldo > 0)
            df_cobr_raw_filtrado,  # raw filtrado (solo clientes con saldo > 0)
            "Cobranza General"     # título
        )
        print(f"  ✅ PDF general → {len(pdf_general):,} bytes")
    except Exception as e:
        pdf_general = None
        print(f"  ⚠️  PDF general: {e}")

print(f"\n✅ df_cobranza: {len(df_cobranza)} clientes")
print(f"✅ PDFs individuales: {len(pdfs_cobranza)}")
df_cobranza.head(5)


# ── bloque-e-compras ──────────────────────────────────────────────────
import math

hoy_str    = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
en_30_dias = (pd.Timestamp.now() + pd.Timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")


# ── 1) Buscar purchase.order confirmados con fecha de entrega próxima ──
po_domain = [
    ("state", "in", ["purchase", "done"]),
    ("receipt_status", "=", "pending"),  # 👈 No recibido
]
po_ids = models.execute_kw(
    ODOO3_DB, uid, ODOO3_PASS,
    "purchase.order", "search",
    [po_domain], {"limit": 10000}
)
print(f"Pedidos de compra encontrados: {len(po_ids)}")

po_list = []
if po_ids:
    po_list = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "purchase.order", "read",
        [po_ids],
        {"fields": ["id", "name", "partner_id", "date_planned", "state"]}
    )

po_map = {p["id"]: p for p in po_list}

# ── 2) Leer líneas de esos pedidos ────────────
pol_ids = models.execute_kw(
    ODOO3_DB, uid, ODOO3_PASS,
    "purchase.order.line", "search",
    [[('order_id', 'in', list(po_map.keys())), ('display_type', '=', False)]],
    {"limit": 300000}
)

pol_list = []
if pol_ids:
    pol_list = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "purchase.order.line", "read",
        [pol_ids],
        {"fields": ["order_id", "product_id", "product_qty", "date_planned"]}
    )

print(f"Líneas de compra encontradas: {len(pol_list)}")

# ── 3) Prefetch productos → categoría padre ───
prd_ids_e = sorted({m2o_id(l.get("product_id")) for l in pol_list if m2o_id(l.get("product_id"))})
prd_to_tmpl_e = {}
if prd_ids_e:
    prods_e = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "product.product", "read", [prd_ids_e], {"fields": ["id", "product_tmpl_id"]}
    )
    prd_to_tmpl_e = {p["id"]: m2o_id(p.get("product_tmpl_id")) for p in prods_e}

tmpl_to_categ_e = {}
tmpl_ids_e = sorted({t for t in prd_to_tmpl_e.values() if t})
if tmpl_ids_e:
    tmpls_e = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "product.template", "read", [tmpl_ids_e], {"fields": ["id", "categ_id"]}
    )
    tmpl_to_categ_e = {t["id"]: m2o_id(t.get("categ_id")) for t in tmpls_e}

# ── Categorías: armar etiqueta "Padre / Hijo" ──
cat_map_e = {}
categ_ids_e = sorted({c for c in tmpl_to_categ_e.values() if c})

if categ_ids_e:
    cats_e = models.execute_kw(
        ODOO3_DB, uid, ODOO3_PASS,
        "product.category", "read", [categ_ids_e],
        {"fields": ["id", "name", "parent_id"]}
    )

    # Guardar hijo + parent_id
    parent_ids_e = set()
    for c in cats_e:
        pid = m2o_id(c.get("parent_id"))
        parent_ids_e.add(pid) if pid else None
        cat_map_e[c["id"]] = {"name": (c.get("name") or ""), "parent_id": pid}

    # Traer los padres que falten
    missing_parents = sorted([pid for pid in parent_ids_e if pid and pid not in cat_map_e])
    if missing_parents:
        pcats_e = models.execute_kw(
            ODOO3_DB, uid, ODOO3_PASS,
            "product.category", "read", [missing_parents],
            {"fields": ["id", "name", "parent_id"]}
        )
        for c in pcats_e:
            cat_map_e[c["id"]] = {"name": (c.get("name") or ""), "parent_id": m2o_id(c.get("parent_id"))}

def cat_padre_hijo(categ_id):
    if not categ_id or categ_id not in cat_map_e:
        return "Sin categoría"

    hijo = (cat_map_e[categ_id].get("name") or "").strip()
    pid  = cat_map_e[categ_id].get("parent_id")

    if pid and pid in cat_map_e:
        padre = (cat_map_e[pid].get("name") or "").strip()
        if padre and hijo:
            return f"{padre} / {hijo}"

    return hijo or "Sin categoría"


# ── 4) Armar DataFrame agrupado por fecha + categoría ─
rows_e = []
for line in pol_list:
    oid  = m2o_id(line.get("order_id"))
    prd  = m2o_id(line.get("product_id"))
    qty  = float(line.get("product_qty") or 0)
    # La fecha puede venir por línea o del pedido
    fecha_str = line.get("date_planned") or (po_map.get(oid) or {}).get("date_planned") or ""
    tmpl_id  = prd_to_tmpl_e.get(prd) if prd else None
    categ_id = tmpl_to_categ_e.get(tmpl_id) if tmpl_id else None
    cat_name = cat_padre_hijo(categ_id)
    rows_e.append({
        "Fecha_entrega": fecha_str,
        "Categoria":     cat_name,
        "Cantidad":      qty,
    })

df_compras = pd.DataFrame(rows_e)
df_compras["Fecha_entrega"] = pd.to_datetime(df_compras["Fecha_entrega"], errors="coerce").dt.normalize()
df_compras["Cantidad"]      = pd.to_numeric(df_compras["Cantidad"], errors="coerce").fillna(0)

# Agrupar por fecha + categoría
df_lo_que_viene = (
    df_compras
    .groupby(["Fecha_entrega", "Categoria"], as_index=False)
    .agg(Unidades=("Cantidad", "sum"))
    .sort_values(["Fecha_entrega", "Categoria"])
)
df_lo_que_viene["Unidades"] = df_lo_que_viene["Unidades"].apply(
    lambda x: int(round(x)) if not math.isnan(x) else 0
)
df_lo_que_viene["Fecha_str"] = df_lo_que_viene["Fecha_entrega"].dt.strftime("%d-%m-%Y")

print(f"\n✅ df_lo_que_viene: {len(df_lo_que_viene)} filas")
df_lo_que_viene[["Fecha_str", "Categoria", "Unidades"]]

# ── bloque-e-html ──────────────────────────────────────────────────
# ── Generar HTML del bloque "Lo que viene" ────

def generar_html_lo_que_viene(df: pd.DataFrame) -> str:
    """Genera párrafos del tipo: 'El 10-02-2026 llegan: Calculadoras (100 u.), Relojes (50 u.)'"""
    if df.empty:
        return "<p>No hay ingresos de mercadería programados para los próximos 30 días.</p>"

    lineas_html = []
    for fecha, grupo in df.groupby("Fecha_str"):
        items = []
        for _, row in grupo.iterrows():
            items.append(f"<b>{row['Categoria']}</b> ({int(row['Unidades'])} u.)")
        lineas_html.append(
            f" <b>{fecha}</b> &nbsp;·&nbsp; {', '.join(items)}</li>"
        )

    return "<ul style='padding-left:18px; margin:6px 0;'>" + "".join(lineas_html) + "</ul>"

html_lo_que_viene = generar_html_lo_que_viene(df_lo_que_viene)

# Vista previa del HTML generado
print(html_lo_que_viene)

# ── c35e4e1e-22d5-427c-b38e-4ddbdabd930a ──────────────────────────────────────────────────
import pandas as pd

def generar_html_lo_que_viene_padre_hijas(df: pd.DataFrame) -> str:
    """
    Espera df con columnas: Fecha_str, Categoria (formato 'Padre / Hijo' o 'Hijo'), Unidades
    Devuelve HTML agrupado por Fecha -> Padre -> Hijas ordenadas
    """
    if df.empty:
        return "<p>No hay ingresos de mercadería programados para los próximos 30 días.</p>"

    t = df.copy()

    # Separar "Padre / Hijo"
    sep = t["Categoria"].astype(str).str.split(r"\s*/\s*", n=1, expand=True)
    t["Padre"] = sep[0].fillna("Sin categoría").replace("", "Sin categoría")
    t["Hijo"]  = sep[1]

    # Si no venía con "Padre / Hijo", tratamos todo como hijo y padre = mismo
    mask_sin_hijo = t["Hijo"].isna() | (t["Hijo"].astype(str).str.strip() == "")
    t.loc[mask_sin_hijo, "Hijo"] = t.loc[mask_sin_hijo, "Padre"]

    # Agrupar por Fecha + Padre + Hijo (por si hay repetidos) y sumar
    t = (t.groupby(["Fecha_str", "Padre", "Hijo"], as_index=False)
           .agg(Unidades=("Unidades", "sum")))

    # Orden
    t["Padre_orden"] = t["Padre"].str.lower()
    t["Hijo_orden"]  = t["Hijo"].str.lower()
    t = t.sort_values(["Fecha_str", "Padre_orden", "Hijo_orden"])

    # Construir HTML
    bloques_fecha = []
    for fecha, df_fecha in t.groupby("Fecha_str", sort=False):
        li_padres = []
        for padre, df_padre in df_fecha.groupby("Padre", sort=False):
            hijas = [
                f"{h} ({int(u)} u.)"
                for h, u in zip(df_padre["Hijo"], df_padre["Unidades"])
                if int(u) != 0
            ]
            if not hijas:
                continue
            li_padres.append(
                f"<li><b>{padre}</b> &nbsp;·&nbsp; " + ", ".join(hijas) + "</li>"
            )

        if li_padres:
            bloques_fecha.append(
                f"<li><b>{fecha}</b><ul style='padding-left:18px; margin:6px 0;'>" +
                "".join(li_padres) +
                "</ul></li>"
            )

    return "<ul style='padding-left:18px; margin:6px 0;'>" + "".join(bloques_fecha) + "</ul>"

html_lo_que_viene = generar_html_lo_que_viene_padre_hijas(df_lo_que_viene)
print(html_lo_que_viene)

# ── bloque-h-resumen ──────────────────────────────────────────────────
RESUMEN_EMAILS = ["natalia@temponovo.cl", "daniel@temponovo.cl"]

def build_email_resumen(df_ventas, df_cobranza, df_cobr_raw,
                        df_diferencias, start, end, html_lo_que_viene):
    rango = (f"{start.strftime('%d-%m-%Y')} al "
             f"{(end-pd.Timedelta(seconds=1)).strftime('%d-%m-%Y')}")

    # ── A) Ventas por familia ──────────────────────────────
    ventas_html = ""
    if not df_ventas.empty:
        try:
            df_v2 = df_ventas.copy()
            def get_familia(ref):
                for pid, p in product_map.items():
                    if (p.get("default_code") or "").strip() == str(ref).strip():
                        tmpl   = prd_to_tmpl_e.get(pid)
                        categ  = tmpl_to_categ_e.get(tmpl) if tmpl else None
                        nombre = cat_padre_hijo(categ) if categ else "Sin categoría"
                        return nombre.split(" / ")[0]
                return "Sin categoría"
            df_v2["Familia"] = df_v2["Referencia"].apply(get_familia)
        except Exception:
            df_v2 = df_ventas.copy()
            df_v2["Familia"] = "Sin categoría"

        fam_grp = (
            df_v2[df_v2["Familia"] != "Sin categoría"]
            .groupby("Familia", as_index=False)
            .agg(Pedidos=("Numero","nunique"),
                 Total=("Total", lambda s: pd.to_numeric(s,errors="coerce").sum()))
            .sort_values("Total", ascending=False)
        )
        # Ventas por vendedor (sin detalle de producto)
        vend_grp = (
            df_ventas.groupby("Vendedor", as_index=False)
            .agg(Pedidos=("Numero","nunique"),
                 Clientes=("Cliente","nunique"),
                 Total=("Total", lambda s: pd.to_numeric(s,errors="coerce").sum()))
            .sort_values("Total", ascending=False)
        )
        total_gral  = pd.to_numeric(df_ventas["Total"], errors="coerce").sum()
        n_ped_total = df_ventas["Numero"].nunique()
        n_cli_total = df_ventas["Cliente"].nunique()

        rows_v = "".join(
            f"<tr><td>{r.Familia}</td>"
            f"<td style='text-align:right'>{r.Pedidos}</td>"
            f"<td style='text-align:right'>$ {format_clp(r.Total)}</td></tr>"
            for r in fam_grp.itertuples()
        )
        rows_vend = "".join(
            f"<tr><td>{r.Vendedor}</td>"
            f"<td style='text-align:right'>{r.Pedidos}</td>"
            f"<td style='text-align:right'>{r.Clientes}</td>"
            f"<td style='text-align:right'>$ {format_clp(r.Total)}</td></tr>"
            for r in vend_grp.itertuples()
        )
        ventas_html = f"""
        <h3>📦 Ventas — por familia</h3>
        <div class="kpi"><b>{n_ped_total} pedidos</b> a <b>{n_cli_total} clientes</b>
        &nbsp;·&nbsp; Total: <b>$ {format_clp(total_gral)}</b></div>
        <table><thead><tr><th>Familia</th>
        <th style='text-align:right'>Pedidos</th>
        <th style='text-align:right'>Total</th></tr></thead>
        <tbody>{rows_v}</tbody></table>
        <h3>📊 Ventas — por vendedor</h3>
        <table><thead><tr><th>Vendedor</th>
        <th style='text-align:right'>Pedidos</th>
        <th style='text-align:right'>Clientes</th>
        <th style='text-align:right'>Total</th></tr></thead>
        <tbody>{rows_vend}</tbody></table>"""

    # ── B) Sin entregar por familia ────────────────────────
    inc_html = ""
    if not df_diferencias.empty:
        try:
            df_inc2 = df_diferencias.copy()
            df_inc2["Familia"] = df_inc2["Referencia"].apply(get_familia)
        except Exception:
            df_inc2 = df_diferencias.copy()
            df_inc2["Familia"] = "Sin categoría"
        df_inc2["Pendiente"] = pd.to_numeric(
            df_inc2.get("Pendiente", pd.Series(dtype=float)), errors="coerce").fillna(0)
        inc_grp = (
            df_inc2[df_inc2["Familia"] != "Sin categoría"]
            .groupby("Familia", as_index=False)
            .agg(Pedidos=("Numero","nunique"), Unidades=("Pendiente","sum"))
            .sort_values("Unidades", ascending=False)
        )
        if not inc_grp.empty:
            rows_i = "".join(
                f"<tr><td>{r.Familia}</td>"
                f"<td style='text-align:right'>{r.Pedidos}</td>"
                f"<td style='text-align:right'>{int(r.Unidades)}</td></tr>"
                for r in inc_grp.itertuples()
            )
            inc_html = f"""
            <h3>⚠️ Sin entregar — por familia</h3>
            <table><thead><tr><th>Familia</th>
            <th style='text-align:right'>Pedidos</th>
            <th style='text-align:right'>Unidades pendientes</th></tr></thead>
            <tbody>{rows_i}</tbody></table>"""

    # ── C) Top 10 clientes más graves ─────────────────────────
    cobr_html = ""
    if not df_cobranza.empty and ">30" in df_cobranza.columns:
        # Totales globales
        total_v30 = float(df_cobranza[">30"].sum())
        total_all = float(df_cobranza["Total"].sum())

        # Top 10: clientes con mayor deuda vencida >30, desempate por antigüedad
        top10 = (
            df_cobranza[df_cobranza[">30"] > 0]
            .copy()
            .assign(
                dias_max = df_cobranza[df_cobranza[">30"] > 0]["Cliente"].map(
                    df_cobr_raw[df_cobr_raw["Bucket"].isin(
                        ["d_31_60","d_61_90","d_91_120","Antiguos"])]
                    .groupby("Cliente")["Dias_vencido"].max()
                )
            )
            .sort_values([">30","dias_max"], ascending=[False, False])
            .head(10)
        )

        def fmt_cell(v, rojo=False):
            v = float(v) if v is not None else 0.0
            if v == 0: return "<td style='text-align:right;color:#ccc'>–</td>"
            st = "color:#c0392b;font-weight:bold" if rojo else ""
            return f"<td style='text-align:right;{st}'>$ {format_clp(v)}</td>"

        # Renombrar columnas con espacios/símbolos para itertuples()
        top10_iter = top10.rename(columns={
            "A la fecha": "A_la_fecha",
            "1-30":       "d_1_30",
            ">30":        "d_mayor30",
        })

        rows_top = ""
        for r in top10_iter.itertuples():
            dias = int(r.dias_max) if hasattr(r,"dias_max") and r.dias_max == r.dias_max else 0
            badge = (f"<span style='background:#c0392b;color:white;border-radius:3px;"
                     f"padding:1px 5px;font-size:11px'>{dias}d</span> "
                     if dias > 0 else "")
            rows_top += (
                f"<tr>"
                f"<td>{badge}{r.Cliente}"
                f"<br><small style='color:#888'>{r.Vendedor} · {r.Ciudad}</small></td>"
                + fmt_cell(float(r.A_la_fecha))
                + fmt_cell(float(r.d_1_30))
                + fmt_cell(float(r.d_mayor30), rojo=True)
                + fmt_cell(float(r.Total))
                + "</tr>"
            )

        cobr_html = f"""
        <h3>💰 Top 10 cobranza más grave</h3>
        <div class="kpi" style="border-color:#c0392b">
          Total pendiente empresa: <b>$ {format_clp(total_all)}</b>
          &nbsp;·&nbsp; Vencido >30: <b style="color:#c0392b">$ {format_clp(total_v30)}</b>
          &nbsp;·&nbsp; <i>PDF adjunto con detalle completo</i>
        </div>
        <table><thead><tr>
          <th>Cliente · Vendedor · Ciudad</th>
          <th style='text-align:right'>A la fecha</th>
          <th style='text-align:right'>1-30</th>
          <th style='text-align:right'>Vencido >30</th>
          <th style='text-align:right'>Total</th>
        </tr></thead>
        <tbody>{rows_top}</tbody></table>
        <p style='font-size:11px;color:#888'>
          El número en rojo indica días de antigüedad máxima del vencimiento.
          Ver PDF adjunto para la cobranza completa ordenada por zona.
        </p>"""

    subject = f"📊 Resumen ejecutivo | {rango}"
    html = f"""
    <html><head>{BASE_CSS}</head><body><div class="wrap">
      <h2>Resumen ejecutivo semanal</h2>
      <p>Semana <b>{rango}</b></p>
      {ventas_html}{inc_html}{cobr_html}
      <div class="lo-que-viene">
        <h3>🔭 Próximos despachos (15 días)</h3>
        {html_lo_que_viene}
      </div>
      <p class="footer">Reporte generado automáticamente desde Odoo.</p>
    </div></body></html>"""
    # ── Excel pedidos: todos los pedidos sin detalle por código ──
    df_pedidos_resumen = pd.DataFrame()
    if not df_ventas.empty:
        df_ped = df_ventas.copy()
        df_ped["Total_num"] = pd.to_numeric(df_ped["Total"], errors="coerce")
        df_ped["Fecha_pedido_str"] = df_ped["Fecha_pedido"].dt.strftime("%d-%m-%Y")             if hasattr(df_ped["Fecha_pedido"], "dt") else df_ped["Fecha_pedido"]
        # Un fila por pedido (Numero único), sin detalle de producto
        df_pedidos_resumen = (
            df_ped.groupby(["Numero","Vendedor","Cliente","Fecha_pedido_str"],
                           as_index=False)
            .agg(Total=("Total_num","sum"))
            .sort_values(["Vendedor","Fecha_pedido_str"])
            .rename(columns={"Numero":"Pedido","Fecha_pedido_str":"Fecha"})
        )
        df_pedidos_resumen["Total"] = df_pedidos_resumen["Total"].apply(format_clp)

    return {"subject": subject, "html": html,
            "df_pedidos": df_pedidos_resumen}

print("✅ build_email_resumen lista")


# ── bloque-f-build ──────────────────────────────────────────────────
def build_email_vendedor(vendedor, dv, df_pendientes, df_cobranza, start, end, html_lo_que_viene):
    nombre     = NOMBRES_CORTOS.get(vendedor, vendedor.title())
    n_ventas   = int(dv["Numero"].nunique())
    n_clientes = int(dv["Cliente"].nunique())
    rango = f"{start.strftime('%d-%m-%Y')} al {(end-pd.Timedelta(seconds=1)).strftime('%d-%m-%Y')}"

    # ── Ventas resumen ──────────────────────────────────────
    ventas_cli = (
        dv.groupby("Cliente", as_index=False)
          .agg(
              Ventas  = ("Numero","nunique"),
              Números = ("Numero", lambda s: ", ".join(sorted(set(map(str,s))))),
              Total   = ("Total",  lambda s: format_clp(pd.to_numeric(s,errors="coerce").sum())),
          ).sort_values("Cliente")
    )

    # ── Diferencias de stock ────────────────────────────────
    dv2 = dv.copy()
    dv2["Entregado"] = pd.to_numeric(dv2["Entregado"],errors="coerce").fillna(0)
    dv2["Pedido"]    = pd.to_numeric(dv2["Pedido"],   errors="coerce").fillna(0)
    flags_v = dv2.assign(hay_cero=(dv2["Entregado"]==0),hay_mayor=(dv2["Entregado"]>0))                 .groupby("Numero",as_index=False)                 .agg(tiene_cero=("hay_cero","any"),tiene_mayor=("hay_mayor","any"))
    nums_inc = flags_v.loc[flags_v["tiene_cero"]&flags_v["tiene_mayor"],"Numero"]
    inc_html = ""
    if not nums_inc.empty:
        inc = dv2[dv2["Numero"].isin(nums_inc)&(dv2["Entregado"]==0)].copy()
        inc["Pendiente"] = (inc["Pedido"]-inc["Entregado"]).clip(lower=0).astype("Int64")
        tbl_inc = inc[["Numero","Cliente","Referencia","Pedido","Entregado","Pendiente"]]                     .sort_values(["Cliente","Numero"])
        inc_html = f"""
          <h3>⚠️ Productos no entregados (stock insuficiente)</h3>
          <p>Estos ítems no se pudieron entregar por falta de stock.</p>
          {df_to_html_table(tbl_inc)}"""

    # ── Pendientes de pago ──────────────────────────────────
    pend_vend = df_pendientes[df_pendientes["Vendedor"].str.strip()==vendedor.strip()].copy()
    pend_html = ""
    if not pend_vend.empty:
        pend_vend["Fecha_pedido"] = pend_vend["Fecha_pedido"].dt.strftime("%d-%m-%Y")
        pend_vend["Total_pedido"] = pend_vend["Total_pedido"].apply(format_clp)
        tbl_pend = pend_vend[["Fecha_pedido","Numero","Cliente","Total_pedido","Categorias"]].copy()
        tbl_pend.columns = ["Fecha","Número","Cliente","Total","Productos"]
        pend_html = f"""
          <h3>⏳ Pedidos esperando pago ({len(tbl_pend)})</h3>
          <p>Estos pedidos están pendientes de pago para salir.</p>
          {df_to_html_table(tbl_pend)}"""

    # ── Cobranza vencida en el email ────────────────────────
    cobr_html = ""
    if not df_cobranza.empty:
        cobr_v = df_cobranza[df_cobranza["Vendedor"].str.strip()==vendedor.strip()]                   .drop(columns=["Vendedor"]).copy()
        if not cobr_v.empty:
            COLS_COBR = ["Cliente","A la fecha","1-30",">30","Total"]
            COLS_VENC = [">30"]

            # Encabezado
            th_cells = "".join(
                f'<th style="width:{"32%" if c=="Cliente" else "9%"};text-align:{"left" if c=="Cliente" else "right"};">{c}</th>'
                for c in COLS_COBR
            )

            def fc(x):
                try:
                    # Convertir a float nativo primero (evita bug con np.float64)
                    v = float(x) if x is not None else 0.0
                    return "" if v == 0 else "$ " + format_clp(v)
                except:
                    return ""

            filas_html = []
            for _, row in cobr_v.iterrows():
              tiene_mayor30 = any(
                  float(str(row.get(c,0) or 0).replace("$ ","").replace(".","").replace(",",".") or 0) > 0
                  for c in COLS_VENC
              )

              tr_class = 'class="fila-vencida"' if tiene_mayor30 else ""

              celdas = []
              for c in COLS_COBR:
                  align = "left" if c == "Cliente" else "right"

                  if c == "Cliente":
                      val = str(row.get("Cliente") or "")
                      extra = ""
                  else:
                      val = fc(row.get(c, 0))
                      extra = ""
                      if c in COLS_VENC and val and tiene_mayor30:
                          extra = ' class="monto-vencido"'
                      elif not val:
                          extra = ' class="vacio"'
                          val = "–"

                  celdas.append(f'<td style="text-align:{align}"{extra}>{val}</td>')

              filas_html.append(f"<tr {tr_class}>{''.join(celdas)}</tr>")


            cobr_html = f"""
              <h3>💰 Cobranza pendiente – por favor revisar</h3>
              <p>Filas en <span style="color:#c0392b;font-weight:bold">rojo</span>
                 = deuda vencida más de 30 días. El PDF adjunto tiene el detalle por factura.</p>
              <table class="tbl-cobr">
                <thead><tr>{th_cells}</tr></thead>
                <tbody>{"".join(filas_html)}</tbody>
              </table>"""

    # ── Excel detalle ───────────────────────────────────────
    dv_excel = dv.copy()
    dv_excel["Fecha_pedido"] = dv_excel["Fecha_pedido"].dt.strftime("%d-%m-%Y")
    cols_exc = ["Fecha_pedido","Numero","Factura","Cliente","Referencia",
                "Pedido","Entregado","Descuento_%","Precio_unit_con_desc","Total"]
    dv_excel = dv_excel[[c for c in cols_exc if c in dv_excel.columns]]

    # ── HTML completo ───────────────────────────────────────
    subject = f"TempoNews – {vendedor} | {rango}"
    html = f"""
    <html><head>{BASE_CSS}</head><body><div class="wrap">
      <h2>Temponovo newsletter</h2>
      <p>Hola <b>{nombre}</b>, espero que estés muy bien,
         te cuento las noticias de esta semana {rango}:</p>

      <h3>📦 Ventas de la semana</h3>
      <div class="kpi">¡Felicidades! Hiciste <b>{n_ventas} ventas</b>
      a <b>{n_clientes} clientes</b></div>
      {df_to_html_table(ventas_cli)}

      {inc_html}
      {pend_html}
      {cobr_html}

      <div class="lo-que-viene">
        <h3>🔭 Próximos despachos (15 días)</h3>
        {html_lo_que_viene}
      </div>

      <p class="footer">Reporte generado automáticamente desde Odoo.<br>
      Ante cualquier duda, comunícate con la oficina.</p>
    </div></body></html>
    """
    return {"subject":subject, "html":html, "detalle_df":dv_excel, "n_ventas":n_ventas}

print("✅ build_email_vendedor lista")


# ── bloque-f-send ──────────────────────────────────────────────────
def send_email(to_email, cc_emails, subject, html_body, attachments=None):
    """attachments = [{data: bytes, filename: str}, ...]"""
    msg = MIMEMultipart()
    msg["From"]    = SMTP_USER
    msg["To"]      = to_email
    msg["Subject"] = subject
    if cc_emails:
        msg["Cc"] = ", ".join(cc_emails)
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    for att in (attachments or []):
        part = MIMEBase("application", "octet-stream")
        part.set_payload(att["data"])
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{att["filename"]}"')
        msg.attach(part)
    recipients = [to_email] + (cc_emails or [])
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, recipients, msg.as_string())

print("✅ Función send_email lista")

# ── bloque-f-run ──────────────────────────────────────────────────
# ═══════════════════════════════════════
# EJECUTAR ENVÍO
# ═══════════════════════════════════════

enviados            = []
omitidos_sin_correo = []
omitidos_sin_ventas = []
tmp_no_borrados     = []

for v in VENDEDORES:
    vendedor = v["name"]

    # Filtrar ventas del vendedor
    dv = df_ventas[df_ventas["Vendedor"].str.strip() == vendedor.strip()].copy() \
         if not df_ventas.empty else pd.DataFrame()

    if dv.empty:
        omitidos_sin_ventas.append(vendedor)
        continue

    # Construir email
    info = build_email_vendedor(vendedor, dv, df_pendientes, df_cobranza, start, end, html_lo_que_viene)

    # Destinatarios
    if TEST_MODE:
        to_email  = TEST_TO[0] if isinstance(TEST_TO, list) else TEST_TO
        cc_emails = TEST_TO[1:] if isinstance(TEST_TO, list) else []
    else:
        to_email = v.get("email")
        if not to_email:
            omitidos_sin_correo.append(vendedor)
            continue
        cc_emails = CC_FIJOS.copy()
        if v.get("email2"):
            cc_emails.append(v["email2"])
        

    # Excel adjunto
    attachments = []
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        export_excel_autowidth(info["detalle_df"], tmp_path, "Detalle ventas")
        with open(tmp_path, "rb") as f:
            attachments.append({
                "data":     f.read(),
                "filename": f"ventas_{safe_filename(vendedor)}_{pd.Timestamp.now().strftime('%Y-%m-%d')}.xlsx"
            })
    finally:
        ok = safe_remove(tmp_path)
        if not ok:
            tmp_no_borrados.append(tmp_path)

    # PDF cobranza
    if vendedor in pdfs_cobranza:
        attachments.append({
            "data":     pdfs_cobranza[vendedor],
            "filename": f"cobranza_{safe_filename(vendedor)}.pdf"
        })

    send_email(to_email, cc_emails, info["subject"], info["html"], attachments)
    modo = "PRUEBA" if TEST_MODE else "REAL"
    print(f"[{modo}] {vendedor} → {to_email} | ventas={info['n_ventas']}")
    enviados.append(vendedor)

# ── Resumen final ─────────────────────────────
print("\n" + "="*55)
print(f"  Enviados       : {', '.join(enviados) or '—'}")
print(f"  Sin correo     : {', '.join(omitidos_sin_correo) or '—'}")
print(f"  Sin ventas     : {', '.join(omitidos_sin_ventas) or '—'}")
if tmp_no_borrados:
    print("  Archivos tmp sin borrar:")
    for p in tmp_no_borrados:
        print(f"    - {p}")
print("="*55)

# ── Resumen ejecutivo ──────────────────────────────────────
try:
    info_res = build_email_resumen(
        df_ventas, df_cobranza, df_cobr_raw,
        df_diferencias, start, end, html_lo_que_viene
    )
    if TEST_MODE:
        to_res = TEST_TO[0] if isinstance(TEST_TO, list) else TEST_TO
        cc_res = TEST_TO[1:] if isinstance(TEST_TO, list) else []
    else:
        to_res = RESUMEN_EMAILS[0]
        cc_res = RESUMEN_EMAILS[1:]
    # Adjuntos del resumen ejecutivo
    attachments_res = []

    # 1) Excel pedidos
    if not info_res["df_pedidos"].empty:
        fd2, tmp_res = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd2)
        try:
            export_excel_autowidth(
                info_res["df_pedidos"], tmp_res, "Pedidos semana"
            )
            with open(tmp_res, "rb") as fx:
                attachments_res.append({
                    "data":     fx.read(),
                    "filename": f"pedidos_semana_{pd.Timestamp.now().strftime('%Y-%m-%d')}.xlsx"
                })
        finally:
            safe_remove(tmp_res)

    # 2) PDF cobranza general (todos los vendedores con zonas)
    if pdf_general:
        attachments_res.append({
            "data":     pdf_general,
            "filename": f"cobranza_general_{pd.Timestamp.now().strftime('%Y-%m-%d')}.pdf"
        })

    send_email(to_res, cc_res, info_res["subject"], info_res["html"], attachments_res)
    modo = "PRUEBA" if TEST_MODE else "REAL"
    print(f"[{modo}] Resumen ejecutivo → {to_res}  (adjunto: {len(attachments_res)} archivos)")
except Exception as e:
    print(f"⚠️  Error enviando resumen: {e}")


# ══════════════════════════════════════════════════════════════
# FIN DEL SCRIPT
# ══════════════════════════════════════════════════════════════
print("✅ Reporte completado exitosamente")
